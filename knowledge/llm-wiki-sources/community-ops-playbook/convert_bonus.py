"""Convert bonus CSV templates to formatted xlsx files matching main template style."""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BONUS_DIR = r"C:\Users\hu\workspace\community-ops-playbook\templates\bonus"
OUTPUT_DIR = r"C:\Users\hu\workspace\community-ops-playbook\templates_xlsx\bonus"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Colors matching main templates
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
INSTRUCTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
INSTRUCTION_FONT = Font(name="Calibri", size=10, italic=True, color="2E7D32")
DATA_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def convert_csv_to_xlsx(csv_path, xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = os.path.splitext(os.path.basename(csv_path))[0]

    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    # Write instruction row (row 1 in original CSV)
    for col_idx, cell_value in enumerate(rows[0], 1):
        cell = ws.cell(row=1, column=col_idx, value=cell_value)
        cell.fill = INSTRUCTION_FILL
        cell.font = INSTRUCTION_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER

    # Write header row (row 2)
    for col_idx, cell_value in enumerate(rows[1], 1):
        cell = ws.cell(row=2, column=col_idx, value=cell_value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
        cell.border = THIN_BORDER

    # Write data rows (row 3+)
    for row_idx, row in enumerate(rows[2:], 3):
        for col_idx, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.font = DATA_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER
            # Alternate row shading
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(rows[0]) + 1):
        max_len = 0
        for row_idx in range(1, min(len(rows) + 1, 25)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 2, 15)

    # Set row heights
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 30

    wb.save(xlsx_path)
    print(f"Created: {xlsx_path}")

csv_files = [f for f in os.listdir(BONUS_DIR) if f.endswith(".csv")]
# Sort to ensure consistent order
csv_files.sort()

for csv_file in csv_files:
    csv_path = os.path.join(BONUS_DIR, csv_file)
    xlsx_file = csv_file.replace(".csv", ".xlsx")
    xlsx_path = os.path.join(OUTPUT_DIR, xlsx_file)
    convert_csv_to_xlsx(csv_path, xlsx_path)

print(f"\nDone. Converted {len(csv_files)} files to {OUTPUT_DIR}")
